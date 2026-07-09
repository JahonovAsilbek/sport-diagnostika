"""Template download — columns = identifying fields + the group's 5 battery exercises (BCKND-58)."""
import io

import pytest
from openpyxl import load_workbook
from rest_framework.test import APIClient

from apps.accounts.factories import UserFactory
from apps.catalog.factories import AgeCategoryFactory
from apps.measurements.import_services import IDENT_COLUMNS
from apps.measurements.tests.import_helpers import TEMPLATE, make_battery

pytestmark = pytest.mark.django_db


def _client(user=None):
    client = APIClient()
    if user is not None:
        client.force_authenticate(user=user)
    return client


def test_template_header_is_ident_plus_battery():
    cat, exercises = make_battery()
    resp = _client(UserFactory(role="super_admin")).get(
        TEMPLATE, {"age_category": cat.id, "gender": "male"}
    )
    assert resp.status_code == 200
    assert resp["Content-Type"].startswith("application/vnd.openxmlformats")
    workbook = load_workbook(io.BytesIO(resp.content))
    header = list(workbook.active.iter_rows(min_row=1, max_row=1, values_only=True))[0]
    assert list(header) == IDENT_COLUMNS + [exercise.name for exercise in exercises]


def test_template_without_battery_is_400():
    cat = AgeCategoryFactory(age_min=7, age_max=8)  # no battery defined
    resp = _client(UserFactory(role="super_admin")).get(
        TEMPLATE, {"age_category": cat.id, "gender": "male"}
    )
    assert resp.status_code == 400


def test_template_unauthenticated_is_401():
    assert _client().get(TEMPLATE).status_code == 401
