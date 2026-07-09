"""Excel import pipeline (B11): template build, per-row validation, and commit. Reuses the
manual-entry services (`parse_raw_value`/`open_session`/`save_measurements`/`finalize_session`)
and `evaluate_session` — scoring is imperative (there is no finalize signal), so commit calls it
explicitly, exactly like the finalize API action."""

from django.conf import settings
from django.db import transaction
from openpyxl import Workbook, load_workbook
from rest_framework.exceptions import ValidationError

from apps.athletes.models import Athlete
from apps.athletes.selectors import AgeOutOfRange, age_category_for
from apps.common.models import Gender
from apps.common.scoping import scope_queryset
from apps.measurements.models import ImportRow, TestSession
from apps.measurements.services import (
    finalize_session,
    open_session,
    parse_raw_value,
    save_measurements,
)
from apps.scoring.domain.battery import battery_for
from apps.scoring.services import evaluate_session

IDENT_COLUMNS = ["last_name", "first_name", "middle_name", "birth_year", "gender"]
_FORMULA_PREFIXES = ("=", "+", "-", "@", "\t", "\r")


class FileImportError(Exception):
    """A file-level problem (unreadable, wrong header, empty, over the row cap) — the whole
    batch fails, distinct from a per-row validation error."""


def sanitize_cell(value):
    """Return a JSON-safe cell value, neutralizing formula-injection: a string cell starting
    with `= + - @` (or a tab/CR) is prefixed with `'` so it can't execute when the stored data
    is later re-exported and opened in the user's spreadsheet app."""
    if value is None or isinstance(value, bool | int | float):
        return value
    text = str(value)
    if text and text[0] in _FORMULA_PREFIXES:
        return "'" + text
    return text


def _text(value):
    return "" if value is None else str(value).strip()


def _int(value):
    try:
        return int(float(value))
    except (TypeError, ValueError):
        return None


def _detail(exc):
    detail = getattr(exc, "detail", exc)
    if isinstance(detail, list) and detail:
        return str(detail[0])
    return str(detail)


def build_template(age_category, gender):
    """An openpyxl `Workbook` whose header is the identifying columns + the group's 5 battery
    exercise names. Raises `ValidationError` if the group has no battery."""
    exercises = battery_for(age_category, gender)
    if exercises is None:
        raise ValidationError("Bu guruh uchun test batareyasi aniqlanmagan.")
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Import"
    worksheet.append(IDENT_COLUMNS + [exercise.name for exercise in exercises])
    return workbook


def read_rows(file_obj, exercises):
    """Parse an uploaded `.xlsx` into `(row_number, {column: sanitized_value})` tuples. Validates
    the header against the template and enforces the row cap. Raises `FileImportError`."""
    try:
        workbook = load_workbook(file_obj, read_only=True, data_only=True)
    except Exception as exc:  # openpyxl raises various errors on a bad file
        raise FileImportError("Fayl o'qib bo'lmadi (yaroqli .xlsx emas).") from exc

    expected = IDENT_COLUMNS + [exercise.name for exercise in exercises]
    rows_iter = workbook.active.iter_rows(values_only=True)
    header = next(rows_iter, None)
    if header is None:
        workbook.close()
        raise FileImportError("Fayl bo'sh.")
    if [_text(cell) for cell in header][: len(expected)] != expected:
        workbook.close()
        raise FileImportError("Fayl sarlavhasi shablonga mos emas.")

    rows = []
    for row_number, row in enumerate(rows_iter, start=2):
        if row is None or all(cell is None or str(cell).strip() == "" for cell in row):
            continue  # skip blank rows
        if len(rows) >= settings.MAX_IMPORT_ROWS:
            workbook.close()
            raise FileImportError(f"Qatorlar soni {settings.MAX_IMPORT_ROWS} dan oshib ketdi.")
        rows.append(
            (
                row_number,
                {col: sanitize_cell(val) for col, val in zip(expected, row, strict=False)},
            )
        )
    workbook.close()
    return rows


def _match_athlete(row_dict, user):
    """Athletes matching the row's natural key within the uploader's scope (up to 2, to detect
    ambiguity). Match-only — imports never create athletes (they're registered via B5)."""
    scoped = scope_queryset(
        Athlete.objects.filter(is_active=True),
        user,
        region_field="region_id",
        organization_field="organization_id",
        coach_field="coach",
    )
    query = scoped.filter(
        last_name__iexact=_text(row_dict.get("last_name")),
        first_name__iexact=_text(row_dict.get("first_name")),
        birth_year=_int(row_dict.get("birth_year")),
        gender=_text(row_dict.get("gender")).lower(),
    )
    middle = _text(row_dict.get("middle_name"))
    if middle:
        query = query.filter(middle_name__iexact=middle)
    return list(query[:2])


def validate_row(row_dict, batch, user, exercises):
    """Validate one parsed row → `(status, errors, athlete)`. Independent — a bad row is never
    fatal to the batch."""
    errors = []
    birth_year = _int(row_dict.get("birth_year"))
    gender = _text(row_dict.get("gender")).lower()
    if not _text(row_dict.get("last_name")) or not _text(row_dict.get("first_name")):
        errors.append({"field": "name", "message": "Familiya va ism majburiy."})
    if birth_year is None:
        errors.append({"field": "birth_year", "message": "Tug'ilgan yil noto'g'ri."})
    if gender not in (Gender.MALE, Gender.FEMALE):
        errors.append({"field": "gender", "message": "Jins male yoki female bo'lishi kerak."})

    athlete = None
    if not errors:
        matches = _match_athlete(row_dict, user)
        if not matches:
            errors.append({"field": "athlete", "message": "Mos sportchi topilmadi."})
        elif len(matches) > 1:
            errors.append({"field": "athlete", "message": "Bir nechta sportchi mos keldi."})
        else:
            athlete = matches[0]

    if athlete is not None:
        # The batch (age_category, gender) is the template group; the session's real TOIFA is
        # recomputed from birth_year — they must agree or the 5 columns aren't this athlete's.
        try:
            if age_category_for(athlete.birth_year, batch.date).id != batch.age_category_id:
                errors.append(
                    {"field": "age_category", "message": "Sportchi TOIFAsi guruhga mos emas."}
                )
        except AgeOutOfRange:
            errors.append(
                {"field": "age_category", "message": "Sportchi yoshi TOIFA oralig'ida emas."}
            )

    for exercise in exercises:
        raw = row_dict.get(exercise.name)
        if raw is None or str(raw).strip() == "":
            errors.append({"field": exercise.name, "message": "Natija majburiy."})
            continue
        try:
            parse_raw_value(exercise.value_type, raw)
        except ValidationError as exc:
            errors.append({"field": exercise.name, "message": _detail(exc)})

    status = ImportRow.Status.ERROR if errors else ImportRow.Status.VALID
    return status, errors, athlete


def commit_batch(batch):
    """Create sessions/measurements/evaluations from the batch's valid rows (each in its own
    savepoint → partial commit). Skips error rows and already-committed rows. Returns the count
    committed. `finalize_session` + `evaluate_session` mirror the manual finalize action."""
    exercises = battery_for(batch.age_category, batch.gender) or []
    by_name = {exercise.name: exercise for exercise in exercises}
    committed = 0
    rows = batch.rows.filter(
        status=ImportRow.Status.VALID, created_session__isnull=True
    ).select_related("athlete")
    for row in rows:
        try:
            with transaction.atomic():
                session = open_session(
                    athlete=row.athlete,
                    entered_by=batch.uploaded_by,
                    date=batch.date,
                    source=TestSession.Source.EXCEL,
                )
                save_measurements(
                    session,
                    [
                        {"exercise": exercise, "raw_value": row.raw_data[name]}
                        for name, exercise in by_name.items()
                    ],
                )
                finalize_session(session)
                evaluate_session(session)
                row.created_session = session
                row.save(update_fields=["created_session", "updated_at"])
            committed += 1
        except Exception as exc:  # one bad row (e.g. missing norm) must not abort the batch
            row.status = ImportRow.Status.ERROR
            row.errors = list(row.errors or []) + [{"field": "commit", "message": _detail(exc)}]
            row.save(update_fields=["status", "errors", "updated_at"])

    batch.status = batch.Status.COMMITTED
    batch.error_count = batch.rows.filter(status=ImportRow.Status.ERROR).count()
    batch.save(update_fields=["status", "error_count", "updated_at"])
    return committed
