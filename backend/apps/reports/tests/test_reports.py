"""Reports — request/202, param scoping, async generation, download 409/file, failure (B12)."""
import io

import pytest
from openpyxl import load_workbook
from rest_framework.test import APIClient

from apps.accounts.factories import UserFactory
from apps.athletes.factories import AthleteFactory
from apps.catalog.factories import OrganizationFactory, RegionFactory
from apps.measurements.factories import TestSessionFactory
from apps.reports.models import Report
from apps.reports.tasks import generate_report
from apps.scoring.factories import EvaluationFactory, IndicatorScoreFactory

pytestmark = pytest.mark.django_db

REPORTS = "/api/v1/reports/"


def _client(user=None):
    client = APIClient()
    if user is not None:
        client.force_authenticate(user=user)
    return client


def _athlete_with_eval(**athlete_kwargs):
    athlete = AthleteFactory(**athlete_kwargs)
    evaluation = EvaluationFactory(session=TestSessionFactory(athlete=athlete), physical_total=42)
    IndicatorScoreFactory(evaluation=evaluation, points=8)
    return athlete


def _download_bytes(response):
    return b"".join(response.streaming_content)


def _require_weasyprint():
    try:
        import weasyprint  # noqa: F401
    except (ImportError, OSError):
        pytest.skip("WeasyPrint / native libs unavailable (Docker-only, DVPS-9)")


# --- request -----------------------------------------------------------------------

def test_unauthenticated_is_401():
    assert _client().post(REPORTS).status_code == 401


def test_request_returns_202_pending():
    athlete = _athlete_with_eval()
    resp = _client(UserFactory(role="super_admin")).post(
        REPORTS,
        {"type": "athlete", "format": "excel", "params": {"athlete": athlete.id}},
        format="json",
    )
    assert resp.status_code == 202
    assert resp.json()["status"] == "pending"


def test_ministry_may_request_a_report():
    resp = _client(UserFactory(role="ministry")).post(
        REPORTS, {"type": "region", "format": "excel", "params": {}}, format="json"
    )
    assert resp.status_code == 202


# --- generation + download ---------------------------------------------------------

def test_excel_athlete_report_generates_and_downloads(django_capture_on_commit_callbacks):
    athlete = _athlete_with_eval()
    client = _client(UserFactory(role="super_admin"))
    with django_capture_on_commit_callbacks(execute=True):
        resp = client.post(
            REPORTS,
            {"type": "athlete", "format": "excel", "params": {"athlete": athlete.id}},
            format="json",
        )
    report_id = resp.json()["id"]
    assert client.get(f"{REPORTS}{report_id}/").json()["status"] == "done"
    download = client.get(f"{REPORTS}{report_id}/download/")
    assert download.status_code == 200
    content = _download_bytes(download)
    assert content[:2] == b"PK"  # xlsx zip
    workbook = load_workbook(io.BytesIO(content))
    assert workbook.active["A1"].value == athlete.full_name


def test_word_report_generates(django_capture_on_commit_callbacks):
    athlete = _athlete_with_eval()
    client = _client(UserFactory(role="super_admin"))
    with django_capture_on_commit_callbacks(execute=True):
        resp = client.post(
            REPORTS,
            {"type": "athlete", "format": "word", "params": {"athlete": athlete.id}},
            format="json",
        )
    assert client.get(f"{REPORTS}{resp.json()['id']}/").json()["status"] == "done"
    content = _download_bytes(client.get(f"{REPORTS}{resp.json()['id']}/download/"))
    assert content[:2] == b"PK"  # docx zip


def test_region_report_lists_the_ranking(django_capture_on_commit_callbacks):
    region = RegionFactory()
    EvaluationFactory(region=region, ranking_score=48, physical_total=48)
    EvaluationFactory(region=region, ranking_score=40, physical_total=40)
    client = _client(UserFactory(role="super_admin"))
    with django_capture_on_commit_callbacks(execute=True):
        resp = client.post(
            REPORTS,
            {"type": "region", "format": "excel", "params": {"region": region.id}},
            format="json",
        )
    report_id = resp.json()["id"]
    content = _download_bytes(client.get(f"{REPORTS}{report_id}/download/"))
    workbook = load_workbook(io.BytesIO(content))
    values = list(workbook.active.iter_rows(values_only=True))
    assert values[0][0] == "Viloyat reytingi"
    assert len(values) >= 4  # title + header + 2 athlete rows


def test_pdf_report_generates_when_weasyprint_available(django_capture_on_commit_callbacks):
    _require_weasyprint()
    athlete = _athlete_with_eval()
    client = _client(UserFactory(role="super_admin"))
    with django_capture_on_commit_callbacks(execute=True):
        resp = client.post(
            REPORTS,
            {"type": "athlete", "format": "pdf", "params": {"athlete": athlete.id}},
            format="json",
        )
    report_id = resp.json()["id"]
    assert client.get(f"{REPORTS}{report_id}/").json()["status"] == "done"
    assert _download_bytes(client.get(f"{REPORTS}{report_id}/download/"))[:4] == b"%PDF"


def test_download_before_done_is_409():
    athlete = _athlete_with_eval()
    client = _client(UserFactory(role="super_admin"))
    resp = client.post(  # no capture → task not run → still pending
        REPORTS,
        {"type": "athlete", "format": "excel", "params": {"athlete": athlete.id}},
        format="json",
    )
    assert client.get(f"{REPORTS}{resp.json()['id']}/download/").status_code == 409


# --- scoping -----------------------------------------------------------------------

def test_region_admin_cannot_request_another_region():
    region_a, region_b = RegionFactory(), RegionFactory()
    admin = UserFactory(role="region_admin", region=region_a)
    resp = _client(admin).post(
        REPORTS, {"type": "region", "format": "excel", "params": {"region": region_b.id}},
        format="json",
    )
    assert resp.status_code == 403


def test_athlete_out_of_scope_is_403():
    region = RegionFactory()
    org = OrganizationFactory(region=region)
    coach = UserFactory(role="coach", organization=org)
    other = _athlete_with_eval(coach=UserFactory(role="coach"), organization=org, region=region)
    resp = _client(coach).post(
        REPORTS, {"type": "athlete", "format": "excel", "params": {"athlete": other.id}},
        format="json",
    )
    assert resp.status_code == 403


def test_report_visible_only_to_requester(django_capture_on_commit_callbacks):
    region = RegionFactory()
    org = OrganizationFactory(region=region)
    coach_a = UserFactory(role="coach", organization=org)
    coach_b = UserFactory(role="coach", organization=org)
    athlete = _athlete_with_eval(coach=coach_a, organization=org, region=region)
    resp = _client(coach_a).post(
        REPORTS, {"type": "athlete", "format": "excel", "params": {"athlete": athlete.id}},
        format="json",
    )
    report_id = resp.json()["id"]
    assert _client(coach_b).get(REPORTS).json()["count"] == 0
    assert _client(coach_b).get(f"{REPORTS}{report_id}/").status_code == 404


# --- failure -----------------------------------------------------------------------

def test_failed_generation_sets_status_failed(monkeypatch):
    athlete = _athlete_with_eval()
    report = Report.objects.create(
        requested_by=UserFactory(role="super_admin"),
        type=Report.Type.ATHLETE, format=Report.Format.EXCEL,
        params={"athlete": athlete.id},
    )
    monkeypatch.setattr(
        "apps.reports.tasks.build_dataset",
        lambda report: (_ for _ in ()).throw(RuntimeError("boom")),
    )
    generate_report(report.id)
    report.refresh_from_db()
    assert report.status == Report.Status.FAILED
    assert "boom" in report.error
    assert report.completed_at is not None
